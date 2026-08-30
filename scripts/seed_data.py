"""RailOpt AI — Frozen Master Dataset Ingestion Script.

Reads ``data/raw/howrah_division_master_data.xlsx`` (READ-ONLY)
and populates PostgreSQL/PostGIS in a single deterministic, repeatable transaction.

IMPORTANT RESEED SEMANTICS:
- The seed pipeline owns source/master/input tables:
    divisions, stations, sections, section_station_map, operational_subsections,
    assets, maintenance_tasks, train_runs, train_section_occupancy, corridor_windows,
    freight_forecasts, resources, resource_availability, block_requests, block_request_tasks.
- The optimizer owns:
    optimization_runs, optimized_blocks, optimized_block_tasks.
- A normal source-data reseed must NEVER delete or mutate optimizer-generated history.

Usage:
    python scripts/seed_data.py
    python scripts/seed_data.py --validate-only
    python scripts/seed_data.py --file path/to/master_data.xlsx
"""

from __future__ import annotations

import argparse
from datetime import date, datetime, time
import logging
from pathlib import Path
import sys
from typing import Any

import openpyxl
from sqlalchemy import create_engine, func, select, text
from sqlalchemy.orm import Session

# Add services/api to sys.path so we can import models and config
PROJECT_ROOT = Path(__file__).resolve().parent.parent
API_DIR = PROJECT_ROOT / "services" / "api"
sys.path.insert(0, str(API_DIR))

from app.core.config import settings
from app.models import (
    Asset,
    Base,
    BlockRequest,
    BlockRequestTask,
    CorridorWindow,
    Division,
    FreightForecast,
    MaintenanceTask,
    OperationalSubsection,
    OptimizationRun,
    OptimizedBlock,
    OptimizedBlockTask,
    Resource,
    ResourceAvailability,
    Section,
    SectionStationMap,
    Station,
    TrainRun,
    TrainSectionOccupancy,
)

# ── Logging Setup ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("railopt.seed")

DEFAULT_WORKBOOK_PATH = (
    PROJECT_ROOT / "data" / "raw" / "howrah_division_master_data.xlsx"
)

VALID_DEPARTMENTS = frozenset({"Engineering", "S&T", "TRD"})
VALID_ASSET_TYPES = frozenset({"Track", "Signal", "OHE"})
VALID_SEVERITIES = frozenset({"Low", "Medium", "High", "Critical"})


# ── Data Parsing Helpers ──────────────────────────────────────────────────────
def clean_str(val: Any) -> str | None:
    """Trim strings and convert empty strings to None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_date(
    val: Any,
    field_name: str = "date",
    sheet_name: str = "",
    row_idx: int | None = None,
) -> date | None:
    """Parse Excel cell value to Python date object."""
    loc = f"[{sheet_name}] Row {row_idx} ({field_name})" if sheet_name else field_name
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        v = val.strip()
        if not v:
            return None
        try:
            return datetime.strptime(v, "%Y-%m-%d").date()
        except ValueError as exc:
            raise ValueError(f"{loc}: Invalid date format '{val}' (expected YYYY-MM-DD)") from exc
    raise ValueError(f"{loc}: Cannot convert type {type(val).__name__} to date: '{val}'")


def parse_time(
    val: Any,
    field_name: str = "time",
    sheet_name: str = "",
    row_idx: int | None = None,
) -> time | None:
    """Parse Excel cell value to Python time object."""
    loc = f"[{sheet_name}] Row {row_idx} ({field_name})" if sheet_name else field_name
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        v = val.strip()
        if not v:
            return None
        # Handle 'YYYY-MM-DD HH:MM' or 'HH:MM' or 'HH:MM:SS'
        if " " in v:
            v_time = v.split(" ")[1]
        else:
            v_time = v
        parts = v_time.split(":")
        try:
            if len(parts) == 2:
                return datetime.strptime(v_time, "%H:%M").time()
            elif len(parts) == 3:
                return datetime.strptime(v_time, "%H:%M:%S").time()
        except ValueError as exc:
            raise ValueError(f"{loc}: Invalid time format '{val}'") from exc
    raise ValueError(f"{loc}: Cannot convert type {type(val).__name__} to time: '{val}'")


def parse_datetime(
    val: Any,
    field_name: str = "datetime",
    sheet_name: str = "",
    row_idx: int | None = None,
) -> datetime | None:
    """Parse Excel cell value to Python datetime object."""
    loc = f"[{sheet_name}] Row {row_idx} ({field_name})" if sheet_name else field_name
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, time.min)
    if isinstance(val, str):
        v = val.strip()
        if not v:
            return None
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(v, fmt)
            except ValueError:
                pass
        raise ValueError(f"{loc}: Invalid datetime format '{val}'")
    raise ValueError(f"{loc}: Cannot convert type {type(val).__name__} to datetime: '{val}'")


def parse_bool(
    val: Any,
    field_name: str = "boolean",
    sheet_name: str = "",
    row_idx: int | None = None,
) -> bool | None:
    """Normalize Excel boolean/string/numeric cell values.

    Accepts:
      - booleans: True, False
      - integers: 0 -> False, >=1 (e.g. platform count) -> True
      - floats: 0.0 -> False, 1.0 -> True
      - strings: 'yes', 'true', '1', 'y' (case-insensitive) -> True
      - strings: 'no', 'false', '0', 'n' (case-insensitive) -> False
      - empty / None -> None

    Fails loudly with ValueError on any unrecognized non-empty value.
    """
    loc = f"[{sheet_name}] Row {row_idx} ({field_name})" if sheet_name else field_name
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, int):
        if val == 1:
            return True
        elif val == 0:
            return False
        elif val > 1:
            # Positive count (e.g. platform_available=24 in stations sheet)
            return True
        elif val < 0:
            return False
    if isinstance(val, float):
        if val == 1.0:
            return True
        elif val == 0.0:
            return False
    if isinstance(val, str):
        v = val.strip().lower()
        if not v:
            return None
        if v in ("yes", "true", "1", "y"):
            return True
        if v in ("no", "false", "0", "n"):
            return False
        raise ValueError(f"{loc}: Unrecognized boolean value: '{val}'")
    raise ValueError(f"{loc}: Cannot convert type {type(val).__name__} with value '{val}' to boolean")


def parse_float(
    val: Any,
    field_name: str = "float",
    sheet_name: str = "",
    row_idx: int | None = None,
) -> float | None:
    """Convert value to float safely."""
    loc = f"[{sheet_name}] Row {row_idx} ({field_name})" if sheet_name else field_name
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        v = val.strip()
        if not v:
            return None
        try:
            return float(v)
        except ValueError as exc:
            raise ValueError(f"{loc}: Invalid float: '{val}'") from exc
    raise ValueError(f"{loc}: Cannot convert type {type(val).__name__} to float: '{val}'")


def parse_int(
    val: Any,
    field_name: str = "int",
    sheet_name: str = "",
    row_idx: int | None = None,
) -> int | None:
    """Convert value to int safely."""
    loc = f"[{sheet_name}] Row {row_idx} ({field_name})" if sheet_name else field_name
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    if isinstance(val, str):
        v = val.strip()
        if not v:
            return None
        try:
            return int(float(v))
        except ValueError as exc:
            raise ValueError(f"{loc}: Invalid integer: '{val}'") from exc
    raise ValueError(f"{loc}: Cannot convert type {type(val).__name__} to int: '{val}'")


# ── Sheet Reader ──────────────────────────────────────────────────────────────
def read_sheet(wb: openpyxl.Workbook, sheet_name: str) -> list[dict[str, Any]]:
    """Read an Excel sheet into a list of row dictionaries."""
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Required sheet '{sheet_name}' not found in workbook!")
    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []
    headers = [str(h).strip() if h is not None else f"col_{idx}" for idx, h in enumerate(header_row)]

    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c is not None for c in row):
            continue  # Skip empty row
        row_dict = {}
        for h, val in zip(headers, row):
            row_dict[h] = val
        row_dict["_row_idx"] = row_idx
        data.append(row_dict)
    return data


# ── Pre-Import Validation ─────────────────────────────────────────────────────
def validate_dataset(wb: openpyxl.Workbook) -> dict[str, list[dict[str, Any]]]:
    """Perform strict pre-import schema and relational validation.

    Fails loudly with detailed error context if any integrity violation is found.
    """
    logger.info("Validating workbook sheets and relationships...")
    sheets = {}

    required_sheets = [
        "01_Division",
        "02_Sections",
        "03_Stations",
        "03B_Section_Station_Map",
        "04_Assets",
        "05_Maintenance_Tasks",
        "06_Train_Schedule",
        "06B_Train_Section_Occupancy",
        "07_Corridor_Windows",
        "08_Freight_Forecast",
        "09_Resources",
        "10_Block_Requests",
    ]

    for s in required_sheets:
        sheets[s] = read_sheet(wb, s)
        logger.info("  Sheet '%s': %d rows found", s, len(sheets[s]))

    # 1. Validate Divisions
    division_names: set[str] = set()
    for row in sheets["01_Division"]:
        d_name = clean_str(row.get("division_name"))
        if not d_name:
            raise ValueError(f"[01_Division] Row {row['_row_idx']}: division_name is required")
        division_names.add(d_name)

    # 2. Validate Stations
    station_codes: set[str] = set()
    for row in sheets["03_Stations"]:
        code = clean_str(row.get("station_code"))
        if not code:
            raise ValueError(f"[03_Stations] Row {row['_row_idx']}: station_code is required")
        if code in station_codes:
            raise ValueError(f"[03_Stations] Duplicate station_code '{code}' at row {row['_row_idx']}")
        station_codes.add(code)

        # Validate boolean fields
        for bool_col in ("block_station", "ibp", "flag_station", "halt", "platform_available", "out_of_division_station"):
            parse_bool(row.get(bool_col), bool_col, "03_Stations", row["_row_idx"])

    # 3. Validate Sections
    section_ids: set[str] = set()
    for row in sheets["02_Sections"]:
        sec_id = clean_str(row.get("section_id"))
        if not sec_id:
            raise ValueError(f"[02_Sections] Row {row['_row_idx']}: section_id is required")
        if sec_id in section_ids:
            raise ValueError(f"[02_Sections] Duplicate section_id '{sec_id}' at row {row['_row_idx']}")
        section_ids.add(sec_id)

        # Validate division reference (Fix 3)
        sec_div = clean_str(row.get("division"))
        if sec_div and sec_div not in division_names:
            raise ValueError(
                f"[02_Sections] Row {row['_row_idx']}: Unknown division '{sec_div}'. "
                f"Known divisions from 01_Division: {list(division_names)}"
            )

        from_st = clean_str(row.get("from_station_code"))
        if from_st and from_st not in station_codes:
            raise ValueError(
                f"[02_Sections] Row {row['_row_idx']}: from_station_code '{from_st}' not found in 03_Stations"
            )
        to_st = clean_str(row.get("to_station_code"))
        if to_st and to_st not in station_codes:
            raise ValueError(
                f"[02_Sections] Row {row['_row_idx']}: to_station_code '{to_st}' not found in 03_Stations"
            )

        # Validate boolean field
        parse_bool(row.get("electrified"), "electrified", "02_Sections", row["_row_idx"])

    # 4. Validate Section-Station Map
    map_pairs: set[tuple[str, str]] = set()
    for row in sheets["03B_Section_Station_Map"]:
        sec_id = clean_str(row.get("section_id"))
        st_code = clean_str(row.get("station_code"))
        if not sec_id or sec_id not in section_ids:
            raise ValueError(f"[03B_Section_Station_Map] Row {row['_row_idx']}: invalid section_id '{sec_id}'")
        if not st_code or st_code not in station_codes:
            raise ValueError(f"[03B_Section_Station_Map] Row {row['_row_idx']}: invalid station_code '{st_code}'")
        pair = (sec_id, st_code)
        if pair in map_pairs:
            raise ValueError(f"[03B_Section_Station_Map] Duplicate mapping pair {pair} at row {row['_row_idx']}")
        map_pairs.add(pair)

    # 5. Validate Assets
    asset_ids: set[str] = set()
    asset_sections: dict[str, str] = {}
    for row in sheets["04_Assets"]:
        a_id = clean_str(row.get("asset_id"))
        if not a_id:
            raise ValueError(f"[04_Assets] Row {row['_row_idx']}: asset_id is required")
        if a_id in asset_ids:
            raise ValueError(f"[04_Assets] Duplicate asset_id '{a_id}' at row {row['_row_idx']}")
        asset_ids.add(a_id)

        sec_id = clean_str(row.get("section_id"))
        if sec_id and sec_id not in section_ids:
            raise ValueError(f"[04_Assets] Row {row['_row_idx']}: section_id '{sec_id}' not found in 02_Sections")
        if sec_id:
            asset_sections[a_id] = sec_id

        st_code = clean_str(row.get("station_code"))
        if st_code and st_code not in station_codes:
            raise ValueError(f"[04_Assets] Row {row['_row_idx']}: station_code '{st_code}' not found in 03_Stations")

        dept = clean_str(row.get("department"))
        if dept not in VALID_DEPARTMENTS:
            raise ValueError(
                f"[04_Assets] Row {row['_row_idx']}: invalid department '{dept}'. Expected {VALID_DEPARTMENTS}"
            )

        asset_type = clean_str(row.get("asset_type"))
        if asset_type not in VALID_ASSET_TYPES:
            raise ValueError(
                f"[04_Assets] Row {row['_row_idx']}: invalid asset_type '{asset_type}'. Expected {VALID_ASSET_TYPES}"
            )

    # 6. Validate Maintenance Tasks
    task_ids: set[str] = set()
    task_sections: dict[str, str] = {}
    for row in sheets["05_Maintenance_Tasks"]:
        t_id = clean_str(row.get("task_id"))
        if not t_id:
            raise ValueError(f"[05_Maintenance_Tasks] Row {row['_row_idx']}: task_id is required")
        if t_id in task_ids:
            raise ValueError(f"[05_Maintenance_Tasks] Duplicate task_id '{t_id}' at row {row['_row_idx']}")
        task_ids.add(t_id)

        a_id = clean_str(row.get("asset_id"))
        if a_id and a_id not in asset_ids:
            raise ValueError(f"[05_Maintenance_Tasks] Row {row['_row_idx']}: asset_id '{a_id}' not found in 04_Assets")

        sec_id = clean_str(row.get("section_id"))
        if sec_id and sec_id not in section_ids:
            raise ValueError(
                f"[05_Maintenance_Tasks] Row {row['_row_idx']}: section_id '{sec_id}' not found in 02_Sections"
            )
        if sec_id:
            task_sections[t_id] = sec_id

        # Validate task.section_id matches asset.section_id
        if a_id and a_id in asset_sections and sec_id:
            if asset_sections[a_id] != sec_id:
                raise ValueError(
                    f"[05_Maintenance_Tasks] Row {row['_row_idx']}: task section '{sec_id}' "
                    f"does not match asset {a_id} section '{asset_sections[a_id]}'"
                )

        dept = clean_str(row.get("dept"))
        if dept not in VALID_DEPARTMENTS:
            raise ValueError(
                f"[05_Maintenance_Tasks] Row {row['_row_idx']}: invalid dept '{dept}'. Expected {VALID_DEPARTMENTS}"
            )

        sev = clean_str(row.get("severity"))
        if sev and sev not in VALID_SEVERITIES:
            raise ValueError(
                f"[05_Maintenance_Tasks] Row {row['_row_idx']}: invalid severity '{sev}'. Expected {VALID_SEVERITIES}"
            )

    # 7. Validate Train Schedule
    run_ids: set[str] = set()
    for row in sheets["06_Train_Schedule"]:
        r_id = clean_str(row.get("run_id"))
        if not r_id:
            raise ValueError(f"[06_Train_Schedule] Row {row['_row_idx']}: run_id is required")
        if r_id in run_ids:
            raise ValueError(f"[06_Train_Schedule] Duplicate run_id '{r_id}' at row {row['_row_idx']}")
        run_ids.add(r_id)

        sec_id = clean_str(row.get("section_id"))
        if sec_id and sec_id not in section_ids:
            raise ValueError(
                f"[06_Train_Schedule] Row {row['_row_idx']}: section_id '{sec_id}' not found in 02_Sections"
            )

        from_code = clean_str(row.get("from_code"))
        if from_code and from_code not in station_codes:
            raise ValueError(
                f"[06_Train_Schedule] Row {row['_row_idx']}: from_code '{from_code}' not found in 03_Stations"
            )

        to_code = clean_str(row.get("to_code"))
        if to_code and to_code not in station_codes:
            raise ValueError(
                f"[06_Train_Schedule] Row {row['_row_idx']}: to_code '{to_code}' not found in 03_Stations"
            )

    # 8. Validate Train Section Occupancy
    occ_ids: set[str] = set()
    for row in sheets["06B_Train_Section_Occupancy"]:
        o_id = clean_str(row.get("occupancy_id"))
        if not o_id:
            raise ValueError(f"[06B_Train_Section_Occupancy] Row {row['_row_idx']}: occupancy_id is required")
        if o_id in occ_ids:
            raise ValueError(f"[06B_Train_Section_Occupancy] Duplicate occupancy_id '{o_id}' at row {row['_row_idx']}")
        occ_ids.add(o_id)

        r_id = clean_str(row.get("run_id"))
        if r_id and r_id not in run_ids:
            raise ValueError(
                f"[06B_Train_Section_Occupancy] Row {row['_row_idx']}: run_id '{r_id}' not found in 06_Train_Schedule"
            )

        sec_id = clean_str(row.get("section_id"))
        if sec_id and sec_id not in section_ids:
            raise ValueError(
                f"[06B_Train_Section_Occupancy] Row {row['_row_idx']}: section_id '{sec_id}' not found in 02_Sections"
            )

    # 9. Validate Corridor Windows
    cw_ids: set[str] = set()
    for row in sheets["07_Corridor_Windows"]:
        w_id = clean_str(row.get("window_id"))
        if not w_id:
            raise ValueError(f"[07_Corridor_Windows] Row {row['_row_idx']}: window_id is required")
        if w_id in cw_ids:
            raise ValueError(f"[07_Corridor_Windows] Duplicate window_id '{w_id}' at row {row['_row_idx']}")
        cw_ids.add(w_id)

        sec_id = clean_str(row.get("section_id"))
        if sec_id and sec_id not in section_ids:
            raise ValueError(
                f"[07_Corridor_Windows] Row {row['_row_idx']}: section_id '{sec_id}' not found in 02_Sections"
            )

    # 10. Validate Freight Forecasts
    ff_pairs: set[tuple[str, Any]] = set()
    for row in sheets["08_Freight_Forecast"]:
        sec_id = clean_str(row.get("section_id"))
        if not sec_id or sec_id not in section_ids:
            raise ValueError(
                f"[08_Freight_Forecast] Row {row['_row_idx']}: invalid section_id '{sec_id}'"
            )
        f_date = parse_date(row.get("date"), "date", "08_Freight_Forecast", row["_row_idx"])
        pair = (sec_id, f_date)
        if pair in ff_pairs:
            raise ValueError(f"[08_Freight_Forecast] Duplicate (section_id, date) {pair} at row {row['_row_idx']}")
        ff_pairs.add(pair)

    # 11. Validate Resources
    res_ids: set[str] = set()
    for row in sheets["09_Resources"]:
        res_id = clean_str(row.get("resource_id"))
        if not res_id:
            raise ValueError(f"[09_Resources] Row {row['_row_idx']}: resource_id is required")
        if res_id in res_ids:
            raise ValueError(f"[09_Resources] Duplicate resource_id '{res_id}' at row {row['_row_idx']}")
        res_ids.add(res_id)

        dept = clean_str(row.get("department"))
        if dept not in VALID_DEPARTMENTS:
            raise ValueError(
                f"[09_Resources] Row {row['_row_idx']}: invalid department '{dept}'. Expected {VALID_DEPARTMENTS}"
            )

    # 12. Validate Block Requests
    req_task_pairs: set[tuple[str, str]] = set()
    for row in sheets["10_Block_Requests"]:
        req_id = clean_str(row.get("request_id"))
        t_id = clean_str(row.get("task_id"))
        if not req_id:
            raise ValueError(f"[10_Block_Requests] Row {row['_row_idx']}: request_id is required")
        if t_id and t_id not in task_ids:
            raise ValueError(
                f"[10_Block_Requests] Row {row['_row_idx']}: task_id '{t_id}' not found in 05_Maintenance_Tasks"
            )
        sec_id = clean_str(row.get("section_id"))
        if sec_id and sec_id not in section_ids:
            raise ValueError(
                f"[10_Block_Requests] Row {row['_row_idx']}: section_id '{sec_id}' not found in 02_Sections"
            )
        dept = clean_str(row.get("dept"))
        if dept and dept not in VALID_DEPARTMENTS:
            raise ValueError(
                f"[10_Block_Requests] Row {row['_row_idx']}: invalid dept '{dept}'. Expected {VALID_DEPARTMENTS}"
            )
        if req_id and t_id:
            pair = (req_id, t_id)
            if pair in req_task_pairs:
                raise ValueError(f"[10_Block_Requests] Duplicate pair {pair} at row {row['_row_idx']}")
            req_task_pairs.add(pair)

    logger.info("Pre-import validation PASSED cleanly across all sheets!")
    return sheets


# ── Database Ingestion ────────────────────────────────────────────────────────
def seed_database(sheets: dict[str, list[dict[str, Any]]], engine) -> dict[str, int]:
    """Insert validated sheet rows into PostgreSQL in strict dependency order.

    All insertions happen within a single database transaction.
    Optimizer-generated tables (optimization_runs, optimized_blocks, optimized_block_tasks)
    are strictly preserved and NEVER deleted.
    """
    logger.info("Starting transactional database seeding...")
    inserted_counts: dict[str, int] = {}

    with Session(engine) as session:
        with session.begin():
            # 1. Clean existing records in reverse dependency order
            # NOTE (Fix 1): Only clear source-owned / master data tables.
            # Optimizer-owned tables (optimization_runs, optimized_blocks, optimized_block_tasks)
            # are deliberately NOT deleted here to preserve optimizer history.
            logger.info("Clearing source-owned master records (reverse dependency order)...")
            session.execute(text("DELETE FROM block_request_tasks;"))
            session.execute(text("DELETE FROM block_requests;"))
            session.execute(text("DELETE FROM corridor_windows;"))
            session.execute(text("DELETE FROM freight_forecasts;"))
            session.execute(text("DELETE FROM train_section_occupancy;"))
            session.execute(text("DELETE FROM train_runs;"))
            session.execute(text("DELETE FROM maintenance_tasks;"))
            session.execute(text("DELETE FROM assets;"))
            session.execute(text("DELETE FROM section_station_map;"))
            session.execute(text("DELETE FROM operational_subsections;"))
            session.execute(text("DELETE FROM sections;"))
            session.execute(text("DELETE FROM stations;"))
            session.execute(text("DELETE FROM resource_availability;"))
            session.execute(text("DELETE FROM resources;"))
            session.execute(text("DELETE FROM divisions;"))

            # 2. Insert Divisions
            logger.info("Inserting divisions...")
            div_map: dict[str, int] = {}
            for row in sheets["01_Division"]:
                div = Division(
                    division_name=clean_str(row.get("division_name")) or "Howrah",
                    zone=clean_str(row.get("zone")) or "Eastern Railway",
                    total_route_km=parse_float(row.get("total_route_km"), "total_route_km", "01_Division", row["_row_idx"]),
                    total_stations=parse_int(row.get("total_stations"), "total_stations", "01_Division", row["_row_idx"]),
                    total_block_stations=parse_int(row.get("total_block_stations"), "total_block_stations", "01_Division", row["_row_idx"]),
                    source=clean_str(row.get("source")),
                    source_type=clean_str(row.get("source_type")),
                )
                session.add(div)
                session.flush()  # To populate div.id
                div_map[div.division_name] = div.id
            inserted_counts["divisions"] = len(sheets["01_Division"])

            # 3. Insert Stations
            logger.info("Inserting stations...")
            for row in sheets["03_Stations"]:
                st = Station(
                    station_code=clean_str(row.get("station_code")) or "",
                    station_name=clean_str(row.get("station_name")) or "",
                    station_type=clean_str(row.get("station_type")),
                    block_station=parse_bool(row.get("block_station"), "block_station", "03_Stations", row["_row_idx"]),
                    ibp=parse_bool(row.get("ibp"), "ibp", "03_Stations", row["_row_idx"]),
                    flag_station=parse_bool(row.get("flag_station"), "flag_station", "03_Stations", row["_row_idx"]),
                    halt=parse_bool(row.get("halt"), "halt", "03_Stations", row["_row_idx"]),
                    platform_available=parse_bool(row.get("platform_available"), "platform_available", "03_Stations", row["_row_idx"]),
                    latitude=parse_float(row.get("latitude"), "latitude", "03_Stations", row["_row_idx"]),
                    longitude=parse_float(row.get("longitude"), "longitude", "03_Stations", row["_row_idx"]),
                    division=clean_str(row.get("division")),
                    zone=clean_str(row.get("zone")),
                    out_of_division_station=parse_bool(row.get("out_of_division_station"), "out_of_division_station", "03_Stations", row["_row_idx"]) or False,
                    administrative_division=clean_str(row.get("administrative_division")),
                    scope_note=clean_str(row.get("scope_note")),
                    source_type=clean_str(row.get("source_type")),
                )
                session.add(st)
            session.flush()
            inserted_counts["stations"] = len(sheets["03_Stations"])

            # 4. Insert Sections (Fix 3: Fail on unknown division)
            logger.info("Inserting sections...")
            for row in sheets["02_Sections"]:
                sec_div_name = clean_str(row.get("division"))
                if sec_div_name:
                    if sec_div_name not in div_map:
                        raise ValueError(
                            f"Unknown division '{sec_div_name}' in 02_Sections row {row['_row_idx']}. "
                            f"Known divisions: {list(div_map.keys())}"
                        )
                    div_id = div_map[sec_div_name]
                else:
                    div_id = None

                sec = Section(
                    section_id=clean_str(row.get("section_id")) or "",
                    section_name=clean_str(row.get("section_name")) or "",
                    from_station_code=clean_str(row.get("from_station_code")),
                    to_station_code=clean_str(row.get("to_station_code")),
                    route_km=parse_float(row.get("route_km"), "route_km", "02_Sections", row["_row_idx"]),
                    track_count=parse_int(row.get("track_count"), "track_count", "02_Sections", row["_row_idx"]),
                    line_type=clean_str(row.get("line_type")),
                    electrified=parse_bool(row.get("electrified"), "electrified", "02_Sections", row["_row_idx"]),
                    signalling_system=clean_str(row.get("signalling_system")),
                    division_id=div_id,
                    source=clean_str(row.get("source")),
                    source_type=clean_str(row.get("source_type")),
                )
                session.add(sec)
            session.flush()
            inserted_counts["sections"] = len(sheets["02_Sections"])

            # 5. Insert Section-Station Map
            logger.info("Inserting section-station mappings...")
            for row in sheets["03B_Section_Station_Map"]:
                ss_map = SectionStationMap(
                    section_id=clean_str(row.get("section_id")) or "",
                    station_code=clean_str(row.get("station_code")) or "",
                    station_sequence=parse_int(row.get("station_sequence"), "station_sequence", "03B_Section_Station_Map", row["_row_idx"]),
                    km_from_section_start=parse_float(row.get("km_from_section_start"), "km_from_section_start", "03B_Section_Station_Map", row["_row_idx"]),
                    relationship_type=clean_str(row.get("relationship_type")),
                )
                session.add(ss_map)
            session.flush()
            inserted_counts["section_station_map"] = len(sheets["03B_Section_Station_Map"])

            # 6. Insert Assets
            logger.info("Inserting assets...")
            for row in sheets["04_Assets"]:
                ast = Asset(
                    asset_id=clean_str(row.get("asset_id")) or "",
                    section_id=clean_str(row.get("section_id")),
                    station_code=clean_str(row.get("station_code")),
                    department=clean_str(row.get("department")) or "",
                    asset_type=clean_str(row.get("asset_type")) or "",
                    failure_risk_score=parse_float(row.get("failure_risk_score"), "failure_risk_score", "04_Assets", row["_row_idx"]),
                    criticality_index=parse_float(row.get("criticality_index"), "criticality_index", "04_Assets", row["_row_idx"]),
                    last_maintained_date=parse_date(row.get("last_maintained_date"), "last_maintained_date", "04_Assets", row["_row_idx"]),
                    source_type=clean_str(row.get("source_type")),
                )
                session.add(ast)
            session.flush()
            inserted_counts["assets"] = len(sheets["04_Assets"])

            # 7. Insert Maintenance Tasks
            logger.info("Inserting maintenance tasks...")
            for row in sheets["05_Maintenance_Tasks"]:
                mt = MaintenanceTask(
                    task_id=clean_str(row.get("task_id")) or "",
                    asset_id=clean_str(row.get("asset_id")),
                    section_id=clean_str(row.get("section_id")),
                    department=clean_str(row.get("dept")) or "",
                    defect_type=clean_str(row.get("defect_type")),
                    severity=clean_str(row.get("severity")),
                    reported_date=parse_date(row.get("reported_date"), "reported_date", "05_Maintenance_Tasks", row["_row_idx"]),
                    days_overdue=parse_int(row.get("days_overdue"), "days_overdue", "05_Maintenance_Tasks", row["_row_idx"]),
                    required_duration_hrs=parse_float(row.get("required_duration_hrs"), "required_duration_hrs", "05_Maintenance_Tasks", row["_row_idx"]),
                    postpone_penalty_cost=parse_float(row.get("postpone_penalty_cost"), "postpone_penalty_cost", "05_Maintenance_Tasks", row["_row_idx"]),
                    priority_score=parse_float(row.get("priority_score"), "priority_score", "05_Maintenance_Tasks", row["_row_idx"]),
                    status=clean_str(row.get("status")) or "Open",
                    source_type=clean_str(row.get("source_type")),
                )
                session.add(mt)
            session.flush()
            inserted_counts["maintenance_tasks"] = len(sheets["05_Maintenance_Tasks"])

            # 8. Insert Train Runs
            logger.info("Inserting train runs...")
            for row in sheets["06_Train_Schedule"]:
                tr = TrainRun(
                    run_id=clean_str(row.get("run_id")) or "",
                    train_no=clean_str(row.get("train_no")) or "",
                    train_name=clean_str(row.get("train_name")),
                    train_type=clean_str(row.get("train_type")),
                    section_id=clean_str(row.get("section_id")),
                    from_station_code=clean_str(row.get("from_code")),
                    to_station_code=clean_str(row.get("to_code")),
                    entry_time=parse_time(row.get("entry_time"), "entry_time", "06_Train_Schedule", row["_row_idx"]),
                    exit_time=parse_time(row.get("exit_time"), "exit_time", "06_Train_Schedule", row["_row_idx"]),
                    priority_rank=parse_int(row.get("priority_rank"), "priority_rank", "06_Train_Schedule", row["_row_idx"]),
                    slack_time_window_mins=parse_int(row.get("slack_time_window_mins"), "slack_time_window_mins", "06_Train_Schedule", row["_row_idx"]),
                    source_type=clean_str(row.get("source_type")),
                )
                session.add(tr)
            session.flush()
            inserted_counts["train_runs"] = len(sheets["06_Train_Schedule"])

            # 9. Insert Train Section Occupancy
            logger.info("Inserting train section occupancy...")
            for row in sheets["06B_Train_Section_Occupancy"]:
                tso = TrainSectionOccupancy(
                    occupancy_id=clean_str(row.get("occupancy_id")) or "",
                    run_id=clean_str(row.get("run_id")),
                    train_id=clean_str(row.get("train_id")),
                    section_id=clean_str(row.get("section_id")),
                    track_id=clean_str(row.get("track_id")),
                    direction=clean_str(row.get("direction")),
                    entry_time=parse_time(row.get("entry_time"), "entry_time", "06B_Train_Section_Occupancy", row["_row_idx"]),
                    exit_time=parse_time(row.get("exit_time"), "exit_time", "06B_Train_Section_Occupancy", row["_row_idx"]),
                    train_type=clean_str(row.get("train_type")),
                    priority_rank=parse_int(row.get("priority_rank"), "priority_rank", "06B_Train_Section_Occupancy", row["_row_idx"]),
                    source_type=clean_str(row.get("source_type")),
                )
                session.add(tso)
            session.flush()
            inserted_counts["train_section_occupancy"] = len(sheets["06B_Train_Section_Occupancy"])

            # 10. Insert Corridor Windows
            logger.info("Inserting corridor windows...")
            for row in sheets["07_Corridor_Windows"]:
                cw = CorridorWindow(
                    window_id=clean_str(row.get("window_id")) or "",
                    section_id=clean_str(row.get("section_id")),
                    window_start=parse_datetime(row.get("window_start"), "window_start", "07_Corridor_Windows", row["_row_idx"]),
                    window_end=parse_datetime(row.get("window_end"), "window_end", "07_Corridor_Windows", row["_row_idx"]),
                    duration_mins=parse_int(row.get("duration_mins"), "duration_mins", "07_Corridor_Windows", row["_row_idx"]),
                    window_type=clean_str(row.get("window_type")),
                    window_status=clean_str(row.get("window_status")),
                    freight_level=clean_str(row.get("freight_level")),
                    source_type=clean_str(row.get("source_type")),
                )
                session.add(cw)
            session.flush()
            inserted_counts["corridor_windows"] = len(sheets["07_Corridor_Windows"])

            # 11. Insert Freight Forecasts
            logger.info("Inserting freight forecasts...")
            for row in sheets["08_Freight_Forecast"]:
                ff = FreightForecast(
                    section_id=clean_str(row.get("section_id")),
                    date=parse_date(row.get("date"), "date", "08_Freight_Forecast", row["_row_idx"]),
                    forecast_freight_trains=parse_int(row.get("forecast_freight_trains"), "forecast_freight_trains", "08_Freight_Forecast", row["_row_idx"]),
                    forecast_tonnage=parse_float(row.get("forecast_tonnage"), "forecast_tonnage", "08_Freight_Forecast", row["_row_idx"]),
                    source_type=clean_str(row.get("source_type")),
                    forecast_confidence=parse_float(row.get("forecast_confidence"), "forecast_confidence", "08_Freight_Forecast", row["_row_idx"]),
                    traffic_level=clean_str(row.get("traffic_level")),
                )
                session.add(ff)
            session.flush()
            inserted_counts["freight_forecasts"] = len(sheets["08_Freight_Forecast"])

            # 12. Insert Resources
            logger.info("Inserting resources...")
            for row in sheets["09_Resources"]:
                res = Resource(
                    resource_id=clean_str(row.get("resource_id")) or "",
                    resource_type=clean_str(row.get("resource_type")),
                    depot=clean_str(row.get("depot")),
                    travel_speed_kmph=parse_float(row.get("travel_speed_kmph"), "travel_speed_kmph", "09_Resources", row["_row_idx"]),
                    availability_window=clean_str(row.get("availability_window")),
                    department=clean_str(row.get("department")) or "",
                    availability_from=parse_time(row.get("availability_from"), "availability_from", "09_Resources", row["_row_idx"]),
                    availability_to=parse_time(row.get("availability_to"), "availability_to", "09_Resources", row["_row_idx"]),
                    team_size=parse_int(row.get("team_size"), "team_size", "09_Resources", row["_row_idx"]),
                    required_skill=clean_str(row.get("required_skill")),
                    equipment=clean_str(row.get("equipment")),
                    status=clean_str(row.get("status")) or "Available",
                    source_type=clean_str(row.get("source_type")),
                )
                session.add(res)
            session.flush()
            inserted_counts["resources"] = len(sheets["09_Resources"])

            # 13. Insert Block Requests & Block Request Tasks
            logger.info("Inserting block requests & request tasks...")
            unique_requests: dict[str, BlockRequest] = {}
            request_tasks: list[BlockRequestTask] = []

            for row in sheets["10_Block_Requests"]:
                req_id = clean_str(row.get("request_id")) or ""
                t_id = clean_str(row.get("task_id"))

                if req_id not in unique_requests:
                    br = BlockRequest(
                        request_id=req_id,
                        department=clean_str(row.get("dept")) or "",
                        section_id=clean_str(row.get("section_id")),
                        requested_start=parse_datetime(row.get("requested_start"), "requested_start", "10_Block_Requests", row["_row_idx"]),
                        requested_duration_hrs=parse_float(row.get("requested_duration_hrs"), "requested_duration_hrs", "10_Block_Requests", row["_row_idx"]),
                        status=clean_str(row.get("status")) or "Pending",
                    )
                    unique_requests[req_id] = br
                    session.add(br)

                if t_id:
                    brt = BlockRequestTask(
                        request_id=req_id,
                        task_id=t_id,
                    )
                    request_tasks.append(brt)

            session.flush()
            for brt in request_tasks:
                session.add(brt)
            session.flush()

            inserted_counts["block_requests"] = len(unique_requests)
            inserted_counts["block_request_tasks"] = len(request_tasks)

    logger.info("Database transaction committed successfully!")
    return inserted_counts


# ── Post-Import Validation & QA ───────────────────────────────────────────────
def verify_seeded_database(engine, expected_counts: dict[str, int]) -> None:
    """Execute post-import count checks and relational integrity QA."""
    logger.info("Running post-import verification and relationship QA...")

    with Session(engine) as session:
        # Check counts
        model_counts = {
            "divisions": session.scalar(select(func.count()).select_from(Division)),
            "stations": session.scalar(select(func.count()).select_from(Station)),
            "sections": session.scalar(select(func.count()).select_from(Section)),
            "section_station_map": session.scalar(select(func.count()).select_from(SectionStationMap)),
            "assets": session.scalar(select(func.count()).select_from(Asset)),
            "maintenance_tasks": session.scalar(select(func.count()).select_from(MaintenanceTask)),
            "train_runs": session.scalar(select(func.count()).select_from(TrainRun)),
            "train_section_occupancy": session.scalar(select(func.count()).select_from(TrainSectionOccupancy)),
            "corridor_windows": session.scalar(select(func.count()).select_from(CorridorWindow)),
            "freight_forecasts": session.scalar(select(func.count()).select_from(FreightForecast)),
            "resources": session.scalar(select(func.count()).select_from(Resource)),
            "block_requests": session.scalar(select(func.count()).select_from(BlockRequest)),
            "block_request_tasks": session.scalar(select(func.count()).select_from(BlockRequestTask)),
        }

        print("\n" + "=" * 60)
        print(f"{'Table':<28} | {'Expected':<10} | {'Inserted':<10} | {'Status'}")
        print("-" * 60)
        for tbl, exp_cnt in expected_counts.items():
            act_cnt = model_counts.get(tbl, 0)
            status = "PASS" if exp_cnt == act_cnt else "FAIL"
            print(f"{tbl:<28} | {exp_cnt:<10} | {act_cnt:<10} | {status}")
            if status == "FAIL":
                raise ValueError(f"Count mismatch in {tbl}: expected {exp_cnt}, found {act_cnt}")
        print("=" * 60 + "\n")

        # Relational QA Checks
        logger.info("Executing relational integrity queries...")

        # 1. Maintenance task -> Asset
        orphan_tasks = session.execute(
            text("SELECT t.task_id FROM maintenance_tasks t LEFT JOIN assets a ON t.asset_id = a.asset_id WHERE a.asset_id IS NULL AND t.asset_id IS NOT NULL;")
        ).fetchall()
        assert len(orphan_tasks) == 0, f"Found orphan maintenance tasks: {orphan_tasks}"

        # 2. Maintenance task -> Section
        orphan_sec_tasks = session.execute(
            text("SELECT t.task_id FROM maintenance_tasks t LEFT JOIN sections s ON t.section_id = s.section_id WHERE s.section_id IS NULL AND t.section_id IS NOT NULL;")
        ).fetchall()
        assert len(orphan_sec_tasks) == 0, f"Found orphan tasks by section: {orphan_sec_tasks}"

        # 3. Asset -> Section
        orphan_assets = session.execute(
            text("SELECT a.asset_id FROM assets a LEFT JOIN sections s ON a.section_id = s.section_id WHERE s.section_id IS NULL AND a.section_id IS NOT NULL;")
        ).fetchall()
        assert len(orphan_assets) == 0, f"Found orphan assets: {orphan_assets}"

        # 4. Task.section_id == Asset.section_id
        mismatched_sections = session.execute(
            text("SELECT t.task_id, t.section_id, a.section_id FROM maintenance_tasks t JOIN assets a ON t.asset_id = a.asset_id WHERE t.section_id != a.section_id;")
        ).fetchall()
        assert len(mismatched_sections) == 0, f"Found task-asset section mismatches: {mismatched_sections}"

        # 5. SectionStationMap -> Section & Station
        orphan_map = session.execute(
            text("SELECT m.id FROM section_station_map m LEFT JOIN sections s ON m.section_id = s.section_id LEFT JOIN stations st ON m.station_code = st.station_code WHERE s.section_id IS NULL OR st.station_code IS NULL;")
        ).fetchall()
        assert len(orphan_map) == 0, f"Found orphan section-station maps: {orphan_map}"

        # 6. TrainRun -> Section & Stations
        orphan_trains = session.execute(
            text("SELECT r.run_id FROM train_runs r LEFT JOIN sections s ON r.section_id = s.section_id LEFT JOIN stations fs ON r.from_station_code = fs.station_code LEFT JOIN stations ts ON r.to_station_code = ts.station_code WHERE (r.section_id IS NOT NULL AND s.section_id IS NULL) OR (r.from_station_code IS NOT NULL AND fs.station_code IS NULL) OR (r.to_station_code IS NOT NULL AND ts.station_code IS NULL);")
        ).fetchall()
        assert len(orphan_trains) == 0, f"Found orphan train runs: {orphan_trains}"

        # 7. Train Section Occupancy -> TrainRun & Section
        orphan_occ = session.execute(
            text("SELECT o.occupancy_id FROM train_section_occupancy o LEFT JOIN train_runs r ON o.run_id = r.run_id LEFT JOIN sections s ON o.section_id = s.section_id WHERE (o.run_id IS NOT NULL AND r.run_id IS NULL) OR (o.section_id IS NOT NULL AND s.section_id IS NULL);")
        ).fetchall()
        assert len(orphan_occ) == 0, f"Found orphan occupancy records: {orphan_occ}"

        # 8. CorridorWindow -> Section
        orphan_cw = session.execute(
            text("SELECT w.window_id FROM corridor_windows w LEFT JOIN sections s ON w.section_id = s.section_id WHERE s.section_id IS NULL AND w.section_id IS NOT NULL;")
        ).fetchall()
        assert len(orphan_cw) == 0, f"Found orphan corridor windows: {orphan_cw}"

        # 9. FreightForecast -> Section
        orphan_ff = session.execute(
            text("SELECT f.id FROM freight_forecasts f LEFT JOIN sections s ON f.section_id = s.section_id WHERE s.section_id IS NULL AND f.section_id IS NOT NULL;")
        ).fetchall()
        assert len(orphan_ff) == 0, f"Found orphan freight forecasts: {orphan_ff}"

        # 10. BlockRequestTask -> BlockRequest & MaintenanceTask
        orphan_brt = session.execute(
            text("SELECT brt.id FROM block_request_tasks brt LEFT JOIN block_requests br ON brt.request_id = br.request_id LEFT JOIN maintenance_tasks mt ON brt.task_id = mt.task_id WHERE br.request_id IS NULL OR mt.task_id IS NULL;")
        ).fetchall()
        assert len(orphan_brt) == 0, f"Found orphan block request tasks: {orphan_brt}"

        logger.info("All 10 relational integrity QA checks PASSED perfectly!")


# ── Main Entrypoint ───────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(description="RailOpt AI Master Dataset Ingestion")
    parser.add_argument(
        "--file",
        type=Path,
        default=DEFAULT_WORKBOOK_PATH,
        help="Path to the frozen master Excel file",
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Run schema and relational validation without modifying the database",
    )
    args = parser.parse_args()

    workbook_path = args.file
    if not workbook_path.is_file():
        logger.error("Master dataset file not found at: %s", workbook_path)
        sys.exit(1)

    logger.info("=" * 60)
    logger.info("RAILOPT AI — MASTER DATA INGESTION PIPELINE")
    logger.info("=" * 60)
    logger.info("Source Workbook (READ-ONLY): %s", workbook_path)
    logger.info("Target Database: %s", settings.database_url.split("@")[-1] if "@" in settings.database_url else "configured")

    # 1. Load Workbook (data_only=True ensures calculated values are read)
    logger.info("Loading Excel workbook...")
    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)

    try:
        # 2. Validate
        sheets = validate_dataset(wb)

        if args.validate_only:
            logger.info("Validation-only flag set. Skipping database insertion.")
            return

        # 3. Seed Database
        engine = create_engine(settings.database_url)
        inserted_counts = seed_database(sheets, engine)

        # 4. Verify & QA
        verify_seeded_database(engine, inserted_counts)

        logger.info("=" * 60)
        logger.info("SEED INGESTION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)

    except Exception as exc:
        logger.error("INGESTION FAILED: %s", exc, exc_info=True)
        sys.exit(1)
    finally:
        wb.close()


if __name__ == "__main__":
    main()
